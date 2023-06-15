from openpyxl import load_workbook, Workbook
import os

# FUNCTION TO ADD DATA IN EXCEL FILE 
def add_data_to_excel(data, filename, sheet_name):
    # CHECK IF THE FILE IS ALREADY EXIST OR NOT 
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name]
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        init_data = ["product_name", "product_description", "product_price", "image_url"]
        for col, name in enumerate(init_data, start=1):
            sheet.cell(row=1, column=col, value=name)
    # SELECT THE LAST ROW
    last_row = sheet.max_row + 1
    # WRITE DATA IN ROWS
    for row, obj in enumerate(data, start=last_row):
        for col, value in enumerate(obj, start=1):
            sheet.cell(row=row, column=col, value=value)
    # SAVE THE FILE
    workbook.save(filename)
